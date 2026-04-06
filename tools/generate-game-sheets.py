"""
Generate Game Stat Sheets
=========================
Reads schedule.json and teams.json, then creates one Excel file per game
inside the 'game-stats/' folder.  Each file has:

  - "Game Info" tab   -> week, date, home/away teams, score fields
  - Home-team tab     -> full roster with blank stat columns
  - Away-team tab     -> full roster with blank stat columns

Stat keepers fill in numbers on game day using their phone / laptop,
then run convert-stats.py to aggregate everything into teams.json.

Usage:
    python tools/generate-game-sheets.py
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---------- paths ----------
ROOT = Path(__file__).resolve().parent.parent
TEAMS_JSON = ROOT / "data" / "teams.json"
SCHEDULE_JSON = ROOT / "data" / "schedule.json"
OUTPUT_DIR = ROOT / "game-stats"

# ---------- stat columns (must match convert-stats.py exactly) ----------
STAT_COLS = [
    "Pass YDs", "Rush YDs", "REC", "Rec YDs",
    "TDs", "INTs", "Sacks", "Flag Pulls"
]

# ---------- styling ----------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
INFO_LABEL_FONT = Font(name="Calibri", bold=True, size=12)
INFO_VALUE_FONT = Font(name="Calibri", size=12)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def team_lookup(teams):
    """Return dict  team_id -> team object."""
    return {t["id"]: t for t in teams}


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def add_game_info_sheet(wb, week, date, home_name, away_name):
    ws = wb.active
    ws.title = "Game Info"
    ws.sheet_properties.tabColor = "1F4E79"

    labels = [
        ("Week", week),
        ("Date", date),
        ("Home Team", home_name),
        ("Away Team", away_name),
        ("Home Score", ""),
        ("Away Score", ""),
    ]

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25

    for i, (label, value) in enumerate(labels, start=1):
        a = ws.cell(row=i, column=1, value=label)
        a.font = INFO_LABEL_FONT
        a.border = THIN_BORDER
        b = ws.cell(row=i, column=2, value=value)
        b.font = INFO_VALUE_FONT
        b.border = THIN_BORDER


def add_team_sheet(wb, team):
    ws = wb.create_sheet(title=team["name"])
    ws.sheet_properties.tabColor = "2E75B6"

    # header row
    headers = ["#", "Name", "POS"] + STAT_COLS
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # column widths
    ws.column_dimensions["A"].width = 5    # #
    ws.column_dimensions["B"].width = 24   # Name
    ws.column_dimensions["C"].width = 6    # POS
    for i in range(len(STAT_COLS)):
        col_letter = chr(ord("D") + i)
        ws.column_dimensions[col_letter].width = 12

    # player rows
    for r, player in enumerate(team["roster"], start=2):
        ws.cell(row=r, column=1, value=player["number"]).border = THIN_BORDER
        ws.cell(row=r, column=2, value=player["name"]).border = THIN_BORDER
        ws.cell(row=r, column=3, value=player["position"]).border = THIN_BORDER
        # stat cells left blank (0 default) — stat keepers fill these in
        for c in range(4, 4 + len(STAT_COLS)):
            cell = ws.cell(row=r, column=c, value=0)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")


def sanitize_filename(name):
    return name.replace(" ", "-").replace("'", "")


def main():
    teams_list = load_json(TEAMS_JSON)
    schedule = load_json(SCHEDULE_JSON)
    teams = team_lookup(teams_list)

    OUTPUT_DIR.mkdir(exist_ok=True)

    count = 0
    for week_data in schedule:
        week = week_data["week"]
        date = week_data["date"]

        for game in week_data["games"]:
            home_id = game["home"]
            away_id = game["away"]
            home_team = teams[home_id]
            away_team = teams[away_id]

            wb = Workbook()
            add_game_info_sheet(wb, week, date, home_team["name"], away_team["name"])
            add_team_sheet(wb, home_team)
            add_team_sheet(wb, away_team)

            filename = f"Week{week}_{sanitize_filename(home_team['name'])}_vs_{sanitize_filename(away_team['name'])}.xlsx"
            filepath = OUTPUT_DIR / filename
            wb.save(filepath)
            count += 1
            print(f"  Created: {filename}")

    print(f"\nDone — {count} game sheets generated in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
