"""
Simulate Week 1 — populate all 7 game sheets with realistic test data.
Run this, then run convert-stats.py to update the JSON files.
"""

import random
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
GAME_STATS_DIR = ROOT / "game-stats"

WEEK1_FILES = sorted(GAME_STATS_DIR.glob("Week1_*.xlsx"))

# Stat columns in order: Pass YDs, Rush YDs, REC, Rec YDs, TDs, INTs, Sacks, Flag Pulls
# Column indices in each team sheet (0=#, 1=Name, 2=POS, 3=Pass YDs, ...)

def generate_player_stats(position):
    """Generate realistic stats based on position."""
    stats = {
        "pass_yds": 0,
        "rush_yds": 0,
        "rec": 0,
        "rec_yds": 0,
        "tds": 0,
        "ints": 0,
        "sacks": 0,
        "flag_pulls": 0,
    }

    if position == "QB":
        stats["pass_yds"] = random.choice([120, 150, 175, 200, 225, 250, 280, 310])
        stats["rush_yds"] = random.choice([0, 5, 10, 15, 20, 30, 45])
        stats["tds"] = random.choice([1, 1, 2, 2, 2, 3, 3, 4])
        stats["ints"] = random.choice([0, 0, 0, 0, 1, 1, 2])
    elif position == "WR":
        stats["rec"] = random.choice([2, 3, 3, 4, 5, 6, 7])
        stats["rec_yds"] = stats["rec"] * random.randint(8, 18)
        stats["rush_yds"] = random.choice([0, 0, 0, 5, 10, 15])
        stats["tds"] = random.choice([0, 0, 0, 1, 1, 1, 2])
        stats["flag_pulls"] = random.choice([0, 0, 1, 1, 2])
    elif position == "RB":
        stats["rush_yds"] = random.choice([25, 35, 45, 55, 65, 80, 90])
        stats["rec"] = random.choice([0, 1, 1, 2, 3])
        stats["rec_yds"] = stats["rec"] * random.randint(5, 12)
        stats["tds"] = random.choice([0, 0, 1, 1, 1, 2])
        stats["flag_pulls"] = random.choice([0, 0, 1, 1])
    elif position in ("S", "DB", "CB", "FS", "SS"):
        stats["ints"] = random.choice([0, 0, 0, 0, 1, 1, 2])
        stats["flag_pulls"] = random.choice([2, 3, 3, 4, 5, 6, 7])
        stats["sacks"] = random.choice([0, 0, 0, 1])
    elif position in ("LB", "DE", "DL", "EDGE"):
        stats["sacks"] = random.choice([0, 0, 1, 1, 1, 2, 2])
        stats["flag_pulls"] = random.choice([1, 2, 3, 3, 4, 5])
        stats["ints"] = random.choice([0, 0, 0, 0, 1])
    elif position in ("C", "OL"):
        # Linemen rarely have stats
        stats["flag_pulls"] = random.choice([0, 0, 0, 1])
    else:
        # Generic / hybrid
        stats["rec"] = random.choice([0, 1, 2, 3])
        stats["rec_yds"] = stats["rec"] * random.randint(5, 15)
        stats["tds"] = random.choice([0, 0, 0, 1])
        stats["flag_pulls"] = random.choice([0, 1, 2, 3])

    return stats


def fill_team_sheet(ws):
    """Fill in stats for every player on a team sheet."""
    rows = list(ws.iter_rows(min_row=2))  # skip header
    for row in rows:
        if row[0].value is None:
            continue
        pos = str(row[2].value).strip() if row[2].value else ""
        s = generate_player_stats(pos)
        row[3].value = s["pass_yds"]    # Pass YDs
        row[4].value = s["rush_yds"]    # Rush YDs
        row[5].value = s["rec"]         # REC
        row[6].value = s["rec_yds"]     # Rec YDs
        row[7].value = s["tds"]         # TDs
        row[8].value = s["ints"]        # INTs
        row[9].value = s["sacks"]       # Sacks
        row[10].value = s["flag_pulls"] # Flag Pulls


def calc_score(ws):
    """Calculate a realistic score from the TDs column (col index 7)."""
    total_tds = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        tds = row[7] if row[7] else 0
        total_tds += tds
    return total_tds * 7  # 7 points per TD (simplified)


for fpath in WEEK1_FILES:
    print(f"Simulating: {fpath.name}")
    wb = load_workbook(fpath)

    # Get team sheet names (everything except "Game Info")
    team_sheets = [s for s in wb.sheetnames if s != "Game Info"]

    # Fill stats for both teams
    for sheet_name in team_sheets:
        fill_team_sheet(wb[sheet_name])

    # Calculate scores from TDs
    gi = wb["Game Info"]
    home_name = gi.cell(row=3, column=2).value  # Home Team value
    away_name = gi.cell(row=4, column=2).value  # Away Team value

    home_score = calc_score(wb[team_sheets[0]])
    away_score = calc_score(wb[team_sheets[1]])

    # Avoid too many ties — nudge one score if tied
    if home_score == away_score and random.random() > 0.2:
        if random.random() > 0.5:
            home_score += 7
        else:
            away_score += 7

    gi.cell(row=5, column=2).value = home_score  # Home Score
    gi.cell(row=6, column=2).value = away_score  # Away Score

    print(f"  {team_sheets[0]} {home_score} - {away_score} {team_sheets[1]}")

    wb.save(fpath)
    wb.close()

print("\nAll Week 1 games simulated! Now run: python tools/convert-stats.py")
