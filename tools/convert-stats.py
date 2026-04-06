"""
Convert Game Sheets -> teams.json + schedule.json + games.json
==============================================================
Reads every .xlsx file in 'game-stats/', then:

  1. Builds data/games.json   — per-game box scores (for the schedule detail view)
  2. Updates data/teams.json  — aggregated player stats + team W/L/T + points
  3. Updates data/schedule.json — fills in game scores

Usage:
    python tools/convert-stats.py

Excel layout expected (matches generate-game-sheets.py):
  - "Game Info" tab with rows: Week, Date, Home Team, Away Team,
    Home Score, Away Score
  - One tab per team with columns:
    #  |  Name  |  POS  |  Pass YDs  |  Rush YDs  |  REC  |
    Rec YDs  |  TDs  |  INTs  |  Sacks  |  Flag Pulls
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook

# ---------- paths ----------
ROOT = Path(__file__).resolve().parent.parent
TEAMS_JSON = ROOT / "data" / "teams.json"
SCHEDULE_JSON = ROOT / "data" / "schedule.json"
GAMES_JSON = ROOT / "data" / "games.json"
GAME_STATS_DIR = ROOT / "game-stats"

# ---------- stat column mapping (Excel header -> JSON key) ----------
STAT_MAP = {
    "Pass YDs":   "passYards",
    "Rush YDs":   "rushYards",
    "REC":        "receptions",
    "Rec YDs":    "recYards",
    "TDs":        "touchdowns",
    "INTs":       "interceptions",
    "Sacks":      "sacks",
    "Flag Pulls": "flagPulls",
}

STAT_KEYS = list(STAT_MAP.values())


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"  Saved: {path}")


def read_game_info(ws):
    info = {}
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=2, values_only=True):
        label, value = row
        if label:
            info[label.strip()] = value
    return info


def read_team_stats(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]

    players = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue

        player = {
            "number": int(row[0]) if row[0] else 0,
            "name": str(row[1]).strip() if row[1] else "",
            "position": str(row[2]).strip() if row[2] else "",
        }

        for col_idx, header in enumerate(headers):
            if header in STAT_MAP:
                json_key = STAT_MAP[header]
                val = row[col_idx] if col_idx < len(row) else 0
                player[json_key] = int(val) if val else 0

        players.append(player)

    return players


def process_game_file(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)

    gi_ws = wb["Game Info"]
    info = read_game_info(gi_ws)

    home_name = str(info.get("Home Team", ""))
    away_name = str(info.get("Away Team", ""))
    home_score = info.get("Home Score")
    away_score = info.get("Away Score")

    if home_score is not None and str(home_score).strip() != "":
        home_score = int(home_score)
    else:
        home_score = None
    if away_score is not None and str(away_score).strip() != "":
        away_score = int(away_score)
    else:
        away_score = None

    team_stats = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "Game Info":
            continue
        ws = wb[sheet_name]
        team_stats[sheet_name] = read_team_stats(ws)

    wb.close()

    return {
        "week": int(info.get("Week", 0)),
        "date": str(info.get("Date", "")),
        "home_name": home_name,
        "away_name": away_name,
        "home_score": home_score,
        "away_score": away_score,
        "teams": team_stats,
    }


def main():
    if not GAME_STATS_DIR.exists():
        print(f"Error: {GAME_STATS_DIR} not found. Run generate-game-sheets.py first.")
        sys.exit(1)

    xlsx_files = sorted(GAME_STATS_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print("No .xlsx files found in game-stats/")
        sys.exit(1)

    # Load current data as templates
    teams_list = load_json(TEAMS_JSON)
    schedule = load_json(SCHEDULE_JSON)

    # Build lookups
    name_to_id = {t["name"]: t["id"] for t in teams_list}
    id_to_team = {t["id"]: t for t in teams_list}

    # Reset all stats to zero before aggregating
    for team in teams_list:
        team["wins"] = 0
        team["losses"] = 0
        team["ties"] = 0
        team["pointsFor"] = 0
        team["pointsAgainst"] = 0
        for player in team["roster"]:
            for key in STAT_KEYS:
                player[key] = 0

    # Reset all schedule scores to null
    for week_data in schedule:
        for sched_game in week_data["games"]:
            sched_game["homeScore"] = None
            sched_game["awayScore"] = None

    # Collect per-game box scores for games.json
    all_games = []

    games_processed = 0
    games_with_scores = 0

    for fpath in xlsx_files:
        print(f"  Reading: {fpath.name}")
        game = process_game_file(fpath)

        # Determine if this game has any stats entered (scores or player stats)
        has_scores = game["home_score"] is not None and game["away_score"] is not None
        has_any_stats = False

        # Build the box score entry
        h_id = name_to_id.get(game["home_name"])
        a_id = name_to_id.get(game["away_name"])

        if not h_id or not a_id:
            print(f"    WARNING: Unknown team in {fpath.name}, skipping")
            continue

        home_players = game["teams"].get(game["home_name"], [])
        away_players = game["teams"].get(game["away_name"], [])

        # Aggregate player stats into teams.json
        for team_name, players in game["teams"].items():
            team_id = name_to_id.get(team_name)
            if not team_id:
                continue

            team_obj = id_to_team[team_id]
            roster_lookup = {p["name"]: p for p in team_obj["roster"]}

            for gp in players:
                rp = roster_lookup.get(gp["name"])
                if not rp:
                    print(f"    WARNING: Unknown player '{gp['name']}' on {team_name}, skipping")
                    continue
                for key in STAT_KEYS:
                    val = gp.get(key, 0)
                    rp[key] += val
                    if val > 0:
                        has_any_stats = True

        # Update W/L/T and scores
        if has_scores:
            games_with_scores += 1
            h_team = id_to_team[h_id]
            a_team = id_to_team[a_id]

            h_team["pointsFor"] += game["home_score"]
            h_team["pointsAgainst"] += game["away_score"]
            a_team["pointsFor"] += game["away_score"]
            a_team["pointsAgainst"] += game["home_score"]

            if game["home_score"] > game["away_score"]:
                h_team["wins"] += 1
                a_team["losses"] += 1
            elif game["home_score"] < game["away_score"]:
                h_team["losses"] += 1
                a_team["wins"] += 1
            else:
                h_team["ties"] += 1
                a_team["ties"] += 1

            # Update schedule.json scores
            for week_data in schedule:
                if week_data["week"] == game["week"]:
                    for sched_game in week_data["games"]:
                        if sched_game["home"] == h_id and sched_game["away"] == a_id:
                            sched_game["homeScore"] = game["home_score"]
                            sched_game["awayScore"] = game["away_score"]

        # Only add to games.json if the game has been played (has scores or stats)
        if has_scores or has_any_stats:
            game_entry = {
                "week": game["week"],
                "date": game["date"],
                "home": h_id,
                "away": a_id,
                "homeScore": game["home_score"],
                "awayScore": game["away_score"],
                "homeStats": home_players,
                "awayStats": away_players,
            }
            all_games.append(game_entry)

        games_processed += 1

    # Sort games by week then by home team
    all_games.sort(key=lambda g: (g["week"], g["home"]))

    # Save all outputs
    save_json(TEAMS_JSON, teams_list)
    save_json(SCHEDULE_JSON, schedule)
    save_json(GAMES_JSON, all_games)

    print(f"\nDone — {games_processed} game files processed, {games_with_scores} had scores entered.")
    print(f"  {len(all_games)} game box scores written to games.json")
    print("  teams.json, schedule.json, and games.json have been updated.")


if __name__ == "__main__":
    main()
